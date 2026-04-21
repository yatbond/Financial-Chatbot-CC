"""
Excel workbook parser.

Handles both .xlsx (openpyxl) and .xls (xlrd) files.
Extracts row-level financial records with source traceability.
Does NOT apply DB mappings — that is the normalizer's job.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

log = logging.getLogger(__name__)

# ── Sheet name normalisation ──────────────────────────────────────────────────

SHEET_ALIASES: dict[str, str] = {
    "Financial Status": "Financial Status",
    "Projection": "Projection",
    "Committed Cost": "Committed Cost",
    "Accrual": "Accrual",
    "Cash Flow": "Cash Flow",
    # Known variants
    "Budget": "Projection",
    "Projected Cost": "Projection",
    "Actual Rec'd & Cost": "Accrual",
    "Cashflow": "Cash Flow",
}

STANDARD_SHEETS = {"Financial Status", "Projection", "Committed Cost", "Accrual", "Cash Flow"}

# Financial Status: column index → raw_financial_type (matches financial_type_map.csv)
FS_VALUE_COLUMNS: dict[int, str] = {
    2:  "Budget Tender",
    3:  "Budget 1st Working Budget",
    4:  "Budget Adjustment Cost/variation",
    5:  "Budget Revision as at",
    6:  "Business Plan",
    7:  "Audit Report (WIP)",
    9:  "Projection as at",
    10: "Committed Value / Cost as at",
    13: "Accrual \n(Before Retention) as at",
    14: "Cash Flow Actual received & paid as at",
}

# Monthly sheets: canonical sheet name → raw_financial_type
MONTHLY_RAW_TYPE: dict[str, str] = {
    "Projection":     "Projection as at",
    "Committed Cost": "Committed Value / Cost as at",
    "Accrual":        "Accrual \n(Before Retention) as at",
    "Cash Flow":      "Cash Flow Actual received & paid as at",
}

MONTH_ABBREVS: dict[str, int] = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4,
    "May": 5, "Jun": 6, "Jul": 7, "Aug": 8,
    "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}

# Rows 1–11 are header/metadata; row 12 is the column-header row for both
# Financial Status and monthly sheets.
METADATA_HEADER_ROWS = 11     # 0-indexed: rows 0..10 are metadata
MONTHLY_DATA_START   = 12     # 0-indexed: row 12 is first data row
FS_DATA_START        = 15     # Financial Status has 4 header rows (12–15)


# ── Data structures ───────────────────────────────────────────────────────────

@dataclass
class ReportHeader:
    project_code: str | None = None
    project_name: str | None = None
    report_date: date | None = None
    report_month: int | None = None
    report_year: int | None = None


@dataclass
class ExtractedRow:
    sheet_name: str
    item_code: str | None
    trade: str | None
    raw_financial_type: str
    value: float | None
    report_month: int
    report_year: int
    source_row_number: int    # 1-indexed Excel row number
    source_col: int           # 0-indexed column
    source_cell_ref: str      # e.g. "C17"


@dataclass
class SheetParseResult:
    canonical_name: str
    original_name: str
    rows: list[ExtractedRow] = field(default_factory=list)
    skipped: bool = False
    skipped_reason: str | None = None
    error: str | None = None


@dataclass
class WorkbookParseResult:
    header: ReportHeader
    sheets: list[SheetParseResult] = field(default_factory=list)


# ── Workbook reading (xlsx / xls abstraction) ─────────────────────────────────

def read_workbook_sheets(path: str) -> dict[str, list[list[Any]]]:
    """Return {sheet_name: [[cell_value, ...], ...]} for all sheets."""
    ext = path.lower().rsplit(".", 1)[-1]
    if ext == "xlsx":
        return _read_xlsx(path)
    if ext == "xls":
        return _read_xls(path)
    raise ValueError(f"Unsupported file extension: .{ext}")


def _read_xlsx(path: str) -> dict[str, list[list[Any]]]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result: dict[str, list[list[Any]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        result[name] = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return result


def _read_xls(path: str) -> dict[str, list[list[Any]]]:
    import xlrd
    wb = xlrd.open_workbook(path)
    result: dict[str, list[list[Any]]] = {}
    for name in wb.sheet_names():
        ws = wb.sheet_by_name(name)
        rows = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    row.append(None)
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    row.append(xlrd.xldate_as_datetime(cell.value, wb.datemode))
                else:
                    row.append(cell.value)
            rows.append(row)
        result[name] = rows
    return result


# ── Header extraction ─────────────────────────────────────────────────────────

def extract_report_header(rows: list[list[Any]]) -> ReportHeader:
    """
    Extract project metadata from the first ~10 rows.
    Expected layout (1-indexed):
      R3: Project Code: <value>
      R4: Project Name: <value>
      R5: Report Date:  <value>
    """
    header = ReportHeader()

    for i, row in enumerate(rows[:11]):
        if not row:
            continue
        label = str(row[0] or "").strip().lower()
        value = row[1] if len(row) > 1 else None

        if "project code" in label and value is not None:
            header.project_code = str(value).strip()
        elif "project name" in label and value is not None:
            header.project_name = str(value).strip()
        elif "report date" in label and value is not None:
            header.report_date = _parse_date(value)

    if header.report_date:
        header.report_month = header.report_date.month
        header.report_year  = header.report_date.year

    return header


def _parse_date(value: Any) -> date | None:
    if isinstance(value, (date, datetime)):
        return value.date() if isinstance(value, datetime) else value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


# ── Item code normalisation ───────────────────────────────────────────────────

def normalize_item_code(raw: Any) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, float):
        # 1.0 → "1", 1.1 → "1.1", 2.12 → "2.12"
        if raw == int(raw):
            return str(int(raw))
        # Avoid floating-point artefacts like "2.1000000000001"
        return re.sub(r"0+$", "", f"{raw:.10f}").rstrip(".")
    s = str(raw).strip()
    return s if s else None


# ── Cell reference helper ─────────────────────────────────────────────────────

def col_letter(col_idx: int) -> str:
    result = ""
    n = col_idx + 1
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def cell_ref(row_idx: int, col_idx: int) -> str:
    """row_idx is 0-indexed; returns e.g. 'C17' (1-indexed row)."""
    return f"{col_letter(col_idx)}{row_idx + 1}"


# ── Financial Status parsing ──────────────────────────────────────────────────

def parse_financial_status(
    rows: list[list[Any]],
    report_month: int,
    report_year: int,
) -> SheetParseResult:
    result = SheetParseResult(canonical_name="Financial Status", original_name="Financial Status")

    for row_idx in range(FS_DATA_START, len(rows)):
        row = rows[row_idx]
        if not row:
            continue

        item_code = normalize_item_code(row[0] if len(row) > 0 else None)
        if item_code is None:
            continue  # skip total/blank rows

        trade = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None

        for col_idx, raw_ft in FS_VALUE_COLUMNS.items():
            val = row[col_idx] if col_idx < len(row) else None
            numeric_val = _to_float(val)

            result.rows.append(ExtractedRow(
                sheet_name="Financial Status",
                item_code=item_code,
                trade=trade,
                raw_financial_type=raw_ft,
                value=numeric_val,
                report_month=report_month,
                report_year=report_year,
                source_row_number=row_idx + 1,
                source_col=col_idx,
                source_cell_ref=cell_ref(row_idx, col_idx),
            ))

    return result


# ── Monthly sheet parsing ─────────────────────────────────────────────────────

def parse_monthly_sheet(
    rows: list[list[Any]],
    canonical_name: str,
    original_name: str,
    report_month: int,
    report_year: int,
) -> SheetParseResult:
    result = SheetParseResult(canonical_name=canonical_name, original_name=original_name)

    raw_ft = MONTHLY_RAW_TYPE.get(canonical_name)
    if not raw_ft:
        result.skipped = True
        result.skipped_reason = f"No raw_financial_type mapping for sheet '{canonical_name}'"
        return result

    # Detect month columns from header row (0-indexed row 11)
    if len(rows) <= METADATA_HEADER_ROWS:
        result.skipped = True
        result.skipped_reason = "Sheet too short to contain a header row"
        return result

    header_row = rows[METADATA_HEADER_ROWS]
    month_cols: dict[int, tuple[int, int]] = {}  # col_idx → (month_num, year)

    for col_idx, cell in enumerate(header_row):
        abbrev = str(cell).strip() if cell is not None else ""
        if abbrev in MONTH_ABBREVS:
            m = MONTH_ABBREVS[abbrev]
            y = _month_year(m, report_month, report_year)
            month_cols[col_idx] = (m, y)

    if not month_cols:
        result.error = "No month columns found in header row"
        return result

    for row_idx in range(MONTHLY_DATA_START, len(rows)):
        row = rows[row_idx]
        if not row:
            continue

        item_code = normalize_item_code(row[0] if len(row) > 0 else None)
        if item_code is None:
            continue

        trade = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None

        for col_idx, (m, y) in month_cols.items():
            val = row[col_idx] if col_idx < len(row) else None
            numeric_val = _to_float(val)

            result.rows.append(ExtractedRow(
                sheet_name=canonical_name,
                item_code=item_code,
                trade=trade,
                raw_financial_type=raw_ft,
                value=numeric_val,
                report_month=m,
                report_year=y,
                source_row_number=row_idx + 1,
                source_col=col_idx,
                source_cell_ref=cell_ref(row_idx, col_idx),
            ))

    return result


# ── Month/year inference ──────────────────────────────────────────────────────

FISCAL_START_MONTH = 4  # April

def _month_year(month: int, report_month: int, report_year: int) -> int:
    """
    Infer the calendar year for a month column given the report date.
    Assumes a fiscal year starting in April (Apr=4).

    For a Feb 2026 report (report_month=2, report_year=2026):
      Apr..Dec → 2025   (prior fiscal year half)
      Jan..Mar → 2026   (current calendar year)
    """
    if report_month >= FISCAL_START_MONTH:
        # Report is in Apr–Dec: fiscal year straddles report_year and report_year+1
        return report_year if month >= FISCAL_START_MONTH else report_year + 1
    else:
        # Report is in Jan–Mar: fiscal year straddles report_year-1 and report_year
        return report_year - 1 if month >= FISCAL_START_MONTH else report_year


# ── Main entry point ──────────────────────────────────────────────────────────

def parse_workbook(path: str) -> WorkbookParseResult:
    """
    Parse an Excel workbook and return extracted rows for all recognised sheets.
    Unrecognised sheets are skipped with a reason logged.
    """
    all_sheet_rows = read_workbook_sheets(path)

    # Use the first recognised sheet to extract the report header
    header = ReportHeader()
    for rows in all_sheet_rows.values():
        h = extract_report_header(rows)
        if h.report_date:
            header = h
            break

    if not header.report_month or not header.report_year:
        log.warning("Could not extract report_month/report_year from workbook header")

    report_month = header.report_month or 1
    report_year  = header.report_year  or 2000

    result = WorkbookParseResult(header=header)

    for original_name, rows in all_sheet_rows.items():
        canonical = SHEET_ALIASES.get(original_name)
        if not canonical:
            log.info("Skipping unrecognised sheet: '%s'", original_name)
            result.sheets.append(SheetParseResult(
                canonical_name=original_name,
                original_name=original_name,
                skipped=True,
                skipped_reason="Sheet name not in recognised list",
            ))
            continue

        log.info("Parsing sheet '%s' (→ '%s')", original_name, canonical)
        try:
            if canonical == "Financial Status":
                sheet_result = parse_financial_status(rows, report_month, report_year)
            else:
                sheet_result = parse_monthly_sheet(rows, canonical, original_name, report_month, report_year)
            sheet_result.original_name = original_name
        except Exception as exc:
            log.exception("Error parsing sheet '%s'", original_name)
            sheet_result = SheetParseResult(
                canonical_name=canonical,
                original_name=original_name,
                error=str(exc),
            )

        result.sheets.append(sheet_result)

    return result


# ── Helpers ───────────────────────────────────────────────────────────────────

def _to_float(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None
